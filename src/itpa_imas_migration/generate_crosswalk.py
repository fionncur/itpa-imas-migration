#!/usr/bin/env python3
"""generate_crosswalk: create a blank crosswalk .xlsx file with IMAS DD lookup formulas.

Writes a two-sheet workbook:
  crosswalk: Crosswalk table; formula columns (imas_unit, imas_dtype,
             imas_description, kind) auto-populate from DD sheet.
             Column A (csv_column) is conditionally coloured by status.
  DD:        IMAS data dictionary extracted from imas-python, as an Excel
             Table named "DD" referenced by the crosswalk formulas.

Pass --from to carry the manual columns (csv_column, status, notes, transform, etc.) over
from an existing crosswalk .xlsx, so a new DD version can be regenerated from a legacy crosswalk.

Usage:
    python generate_crosswalk [--dd-version 4.1.1] [--output path/to/out.xlsx]
                              [--num-rows 200] [--from path/to/existing_crosswalk.xlsx]
"""

import argparse
import pathlib

import imas
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def _find_repo_root(start: pathlib.Path) -> pathlib.Path:
    """Walk up from start to the repo root (the dir holding resources/ or pyproject.toml)."""
    for candidate in (start, *start.parents):
        if (candidate / "resources").is_dir() or (candidate / "pyproject.toml").is_file():
            return candidate
    return start


HERE = pathlib.Path(__file__).resolve().parent
ROOT = _find_repo_root(HERE)

CROSSWALK_COLUMNS = [
    "csv_column",
    "csv_unit",
    "imas_unit",
    "csv_dtype",
    "imas_dtype",
    "imas_path",
    "csv_description",
    "imas_standard_name",
    "imas_description",
    "kind",
    "status",
    "notes",
    "transform",
    "transform_args",
    "needs_source",
    "source_fields",
    "source",
    "errors",
]

DD_COLUMNS = ["identifier", "description", "units", "data_type", "kind"]

# Columns whose values are filled by formulas (not user input).
FORMULA_COLS = {"imas_unit", "imas_dtype", "imas_description", "kind"}

# status → background hex (RRGGBB, no alpha prefix).
STATUS_COLORS = {
    "mapped": "C6DEB5",  # light green
    "mapped_caveat": "FEEE76",  # yellow
    "manifest": "FF5B5B",  # red
    "derived": "8FAADC",  # light blue
    "discard": "B1A0C7",  # light purple
}


def _extract_dd_rows(root) -> list[dict]:
    """Walk the imas-python DD XML tree and yield one dict per identifier."""
    rows: list[dict] = []

    for ids_el in root.findall("IDS"):
        ids_name = ids_el.attrib.get("name", "")
        rows.append(
            {
                "identifier": ids_name,
                "description": ids_el.attrib.get("documentation", ""),
                "units": "",
                "data_type": "IDS",
                "kind": ids_el.attrib.get("type", ""),
            }
        )
        for field_el in ids_el.iter("field"):
            path = field_el.attrib.get("path", "")
            rows.append(
                {
                    "identifier": f"{ids_name}/{path}",
                    "description": field_el.attrib.get("documentation", ""),
                    "units": field_el.attrib.get("units", ""),
                    "data_type": field_el.attrib.get("data_type", ""),
                    "kind": field_el.attrib.get("type", ""),
                }
            )

    return rows


# Column S (hidden) holds the normalised imas_path used by C/E/I/J.
# It strips the first &-delimited path and removes index suffixes in parentheses,
# like (0), (:) etc.
NORM_COL = "S"
NORM_COL_IDX = 19  # 1-based


def _formula_norm(row: int) -> str:
    """Path normalisation in the hidden helper column S,

    1. Take only the first &-separated path.
    2. Strip parenthetical index suffixes.
    """
    p = f"$F{row}"
    first = f'LEFT({p},IFERROR(FIND("&",{p}),LEN({p})+1)-1)'
    stripped = first
    for token in ["(0)", "(1)", "(2)", "(3)", "(4)", "(5)", "(6)", "(7)", "(8)", "(9)", "(:)"]:
        stripped = f'SUBSTITUTE({stripped},"{token}","")'
    return f'=IF({p}="","",{stripped})'


def _formula_imas_unit(row: int) -> str:
    n = f"${NORM_COL}{row}"
    return (
        f'=IFERROR(IF(INDEX(DD[units],MATCH({n},DD[identifier],0))=0,"",'
        f'INDEX(DD[units],MATCH({n},DD[identifier],0))),"")'
    )


def _formula_imas_dtype(row: int) -> str:
    n, ns = f"${NORM_COL}{row}", f"$O{row}"
    return f'=IFERROR(INDEX(DD[data_type],MATCH(IF({ns}=TRUE,{n}&"/value",{n}),' f'DD[identifier],0)),"")'


def _formula_imas_description(row: int) -> str:
    n = f"${NORM_COL}{row}"
    return f'=IFERROR(INDEX(DD[description],MATCH({n},DD[identifier],0)),"")'


def _formula_kind(row: int) -> str:
    n, ns = f"${NORM_COL}{row}", f"$O{row}"
    return f'=IFERROR(INDEX(DD[kind],MATCH(IF({ns},{n}&"/value",{n}),' f'DD[identifier],0)),"")'


FORMULA_BUILDERS = {
    "imas_unit": _formula_imas_unit,
    "imas_dtype": _formula_imas_dtype,
    "imas_description": _formula_imas_description,
    "kind": _formula_kind,
}


def _build_dd_sheet(wb: openpyxl.Workbook, dd_rows: list[dict]) -> None:
    ws = wb.create_sheet("DD")

    # Generate headers
    for col_idx, col_name in enumerate(DD_COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Variable rows: write values, leave empty cells as None.
    for row_idx, row in enumerate(dd_rows, start=2):
        for col_idx, col_name in enumerate(DD_COLUMNS, start=1):
            value = row.get(col_name, "")
            ws.cell(row=row_idx, column=col_idx, value=value if value else None)

    # Excel Table so the crosswalk sheet can reference DD[identifier] etc.
    n_rows = len(dd_rows)
    last_col = get_column_letter(len(DD_COLUMNS))
    tbl = Table(displayName="DD", ref=f"A1:{last_col}{n_rows + 1}")
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight1",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
    )
    ws.add_table(tbl)

    # Column widths
    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 12


def _build_crosswalk_sheet(wb: openpyxl.Workbook, num_rows: int) -> None:
    ws = wb.active
    ws.title = "crosswalk"

    col_map = {name: get_column_letter(i + 1) for i, name in enumerate(CROSSWALK_COLUMNS)}
    status_col_letter = col_map["status"]  # K
    csv_col_letter = col_map["csv_column"]  # A

    # Header row (no formatting)
    for col_idx, col_name in enumerate(CROSSWALK_COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Variable rows: normalised-path helper (col S, hidden) + visible formula columns
    for row in range(2, num_rows + 2):
        ws.cell(row=row, column=NORM_COL_IDX, value=_formula_norm(row))
        for col_name, builder in FORMULA_BUILDERS.items():
            col_idx = CROSSWALK_COLUMNS.index(col_name) + 1
            ws.cell(row=row, column=col_idx, value=builder(row))

    ws.column_dimensions[NORM_COL].hidden = True

    # Conditional formatting: colour csv_column (col A) by status (col K).
    for status, hex_color in STATUS_COLORS.items():
        fill = PatternFill(bgColor="FF" + hex_color)
        dxf = DifferentialStyle(fill=fill)
        formula = f'${status_col_letter}1="{status}"'
        rule = Rule(
            type="expression",
            dxf=dxf,
            formula=[formula],
        )
        ws.conditional_formatting.add(f"{csv_col_letter}1:{csv_col_letter}1048576", rule)

    # Column widths
    widths = {
        "csv_column": 16,
        "csv_unit": 10,
        "imas_unit": 12,
        "csv_dtype": 22,
        "imas_dtype": 16,
        "imas_path": 48,
        "csv_description": 40,
        "imas_standard_name": 32,
        "imas_description": 48,
        "kind": 10,
        "status": 14,
        "notes": 36,
        "transform": 12,
        "transform_args": 36,
        "needs_source": 14,
        "source_fields": 22,
        "source": 22,
        "errors": 30,
    }
    for col_name, width in widths.items():
        ws.column_dimensions[col_map[col_name]].width = width

    # Wrap text in every cell.
    wrap = Alignment(wrap_text=True)
    for row in range(1, num_rows + 2):
        for col_idx in range(1, len(CROSSWALK_COLUMNS) + 1):
            ws.cell(row=row, column=col_idx).alignment = wrap

    # Freeze header row and first column
    ws.freeze_panes = "B2"


MANUAL_COLS = [c for c in CROSSWALK_COLUMNS if c not in FORMULA_COLS]


def _coerce(col_name: str, val) -> object:
    if not isinstance(val, str) and pd.isna(val):
        return None
    if col_name == "needs_source":
        if isinstance(val, str):
            return True if val.strip().upper() == "TRUE" else None
        return True if val else None
    return val


def _copy_manual_data(ws, df: pd.DataFrame) -> None:
    """Write manual columns from an already-loaded crosswalk dataframe into ws."""
    col_idx = {name: CROSSWALK_COLUMNS.index(name) + 1 for name in MANUAL_COLS if name in df.columns}

    for df_row_idx, df_row in enumerate(df.itertuples(index=False), start=2):
        for col_name, col_num in col_idx.items():
            val = _coerce(col_name, getattr(df_row, col_name, None))
            ws.cell(row=df_row_idx, column=col_num, value=val)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate a blank crosswalk .xlsx with IMAS DD lookup formulas.",
    )
    parser.add_argument("--dd-version", default="4.1.1", metavar="VERSION")
    parser.add_argument(
        "--output",
        default=str(ROOT / "resources" / "mappings" / "new_crosswalk.xlsx"),
        metavar="PATH",
    )
    parser.add_argument(
        "--num-rows",
        type=int,
        default=200,
        help="Number of variable rows to pre-populate with formulas (default: 200).",
    )
    parser.add_argument(
        "--from",
        dest="from_existing",
        default=None,
        metavar="PATH",
        help="Existing crosswalk .xlsx to copy manual data from.",
    )
    args = parser.parse_args()

    print(f"Loading IMAS DD version {args.dd_version} ...")
    factory = imas.IDSFactory(args.dd_version)
    root = factory._etree.getroot()

    print("Extracting DD rows ...")
    dd_rows = _extract_dd_rows(root)
    print(f"  {len(dd_rows):,} identifiers found.")

    from_df = None
    num_rows = args.num_rows
    if args.from_existing:
        src = pathlib.Path(args.from_existing)
        print(f"Reading existing crosswalk {src} ...")
        from_df = pd.read_excel(src, sheet_name=0)
        if len(from_df) > num_rows:
            print(f"  {len(from_df)} rows in source exceeds --num-rows={num_rows}; using {len(from_df)}.")
            num_rows = len(from_df)

    wb = openpyxl.Workbook()
    _build_crosswalk_sheet(wb, num_rows=num_rows)
    _build_dd_sheet(wb, dd_rows)

    if from_df is not None:
        print("Copying manual data ...")
        _copy_manual_data(wb["crosswalk"], from_df)

    out = pathlib.Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Saved: {out}")


if __name__ == "__main__":
    main()
